from __future__ import annotations

import html as html_lib
import json
import re
import shutil
import sys
from base64 import b64encode
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path
from statistics import mean, median
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
ASSETS_DIR = ROOT / "assets"
LOGO_PATH = ASSETS_DIR / "alforaij_logo.png"
COVER_PATH = ASSETS_DIR / "kuwait_glass_cover.png"
PLATFORM_DIR = (
    ROOT
    / "output"
    / "للمدير_فقط_حزمة_شاملة_نهائية_المحافظات_والفجوات"
    / "01_إحصائيات_الفريج_من_المنصة"
)
OUTPUT_DIR = PLATFORM_DIR / "05_استعراض_الأرقام_والعروض_الفعلية"
DETAILS_DIR = OUTPUT_DIR / "03_pages"
HTML_PATH = OUTPUT_DIR / "01_لوحة_تفاعلية_للأرقام_والعروض.html"
PROPERTY_HTML_PATH = OUTPUT_DIR / "01_لوحة_تفاعلية_للأرقام_والعروض_نسخة_نوع_العقار.html"
XLSX_PATH = OUTPUT_DIR / "02_قوائم_العروض_الفعلية_حسب_الأرقام.xlsx"
README_PATH = OUTPUT_DIR / "اقرأني_مختصر.txt"
PUBLIC_DETAIL_BASE = "https://front.alforaij.com/Listing/Detail"
PUBLIC_SEARCH_URL = "https://search.alforaij.com/"
IMAGE_BASE = "https://search.alforaij.com"

CORE_TYPES = {"للبيع", "للإيجار", "مطلوب للشراء", "مطلوب للإيجار"}
BRAND = {
    "ink": "111827",
    "muted": "475569",
    "line": "CBD5E1",
    "bg": "F6F8FB",
    "panel": "FFFFFF",
    "navy": "0F172A",
    "blue": "2554D9",
    "orange": "D97706",
    "green": "047857",
    "red": "B91C1C",
}

ARABIC_MONTHS = {
    1: "يناير",
    2: "فبراير",
    3: "مارس",
    4: "أبريل",
    5: "مايو",
    6: "يونيو",
    7: "يوليو",
    8: "أغسطس",
    9: "سبتمبر",
    10: "أكتوبر",
    11: "نوفمبر",
    12: "ديسمبر",
}


def arabic_date_label(value: date) -> str:
    return f"حتى {value.day} {ARABIC_MONTHS[value.month]} {value.year}"


def image_data_uri(path: Path) -> str:
    if not path.exists() or path.stat().st_size == 0:
        return ""
    suffix = path.suffix.lower()
    mime = "image/jpeg" if suffix in {".jpg", ".jpeg"} else "image/png"
    return f"data:{mime};base64,{b64encode(path.read_bytes()).decode('ascii')}"


def find_sheet(wb, title: str, fallback_index: int):
    if title in wb.sheetnames:
        return wb[title]
    return wb.worksheets[fallback_index]


def clean_value(value):
    if value is None:
        return ""
    return value


def as_float(value):
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def fmt_int(value: int | float | None) -> str:
    if value is None:
        return "0"
    return f"{int(round(float(value))):,}"


def fmt_price(value: int | float | None, unit: str = "د.ك") -> str:
    if value is None:
        return "غير معلن"
    number = int(round(float(value)))
    if number >= 10_000:
        return f"{number / 1000:,.0f} ألف {unit}"
    return f"{number:,.0f} {unit}"


def safe_slug(text: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", text)
    slug = re.sub(r"_+", "_", slug).strip("_")
    return slug[:24] or "tbl"


def listing_id_from_code(code: str) -> str:
    match = re.search(r"(\d+)$", str(code or ""))
    return match.group(1) if match else ""


def local_detail_url(code: str) -> str:
    return ""


def public_detail_url(code: str) -> str:
    listing_id = listing_id_from_code(code)
    return f"{PUBLIC_DETAIL_BASE}/{listing_id}" if listing_id else ""


def original_search_url(code: str) -> str:
    return public_detail_url(code)


def public_image_url(path: object) -> str:
    text = str(path or "").strip()
    if not text:
        return ""
    if text.startswith("http://") or text.startswith("https://"):
        return text
    if text.startswith("/"):
        return f"{IMAGE_BASE}{text}"
    return f"{IMAGE_BASE}/{text}"


def clean_detail_text(value: object) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


def fetch_listing_visual(code: str) -> tuple[str, dict[str, str]]:
    listing_id = listing_id_from_code(code)
    if not listing_id:
        return code, {"imageUrl": "", "detailTitle": "", "detailText": ""}
    try:
        req = Request(
            f"https://search.alforaij.com/api/internallistings/{listing_id}",
            headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json,text/plain,*/*"},
        )
        with urlopen(req, timeout=12) as response:
            payload = json.loads(response.read().decode("utf-8"))
        data = payload.get("data", payload) if isinstance(payload, dict) else {}
        if not isinstance(data, dict):
            return code, {"imageUrl": "", "detailTitle": "", "detailText": ""}
        image = public_image_url(data.get("featuredImagePath"))
        if not image:
            raw_images = data.get("imagesJson") or ""
            try:
                parsed = json.loads(raw_images) if raw_images else []
            except json.JSONDecodeError:
                parsed = []
            if parsed:
                image = public_image_url(parsed[0])
        return code, {
            "imageUrl": image,
            "detailTitle": clean_detail_text(data.get("titleAr") or data.get("titleEn")),
            "detailText": clean_detail_text(data.get("descriptionAr") or data.get("descriptionEn")),
        }
    except Exception:
        return code, {"imageUrl": "", "detailTitle": "", "detailText": ""}


def fetch_listing_visuals(codes: list[str]) -> dict[str, dict[str, str]]:
    unique_codes = list(dict.fromkeys(codes))
    visuals: dict[str, dict[str, str]] = {}
    with ThreadPoolExecutor(max_workers=18) as executor:
        for code, visual in executor.map(fetch_listing_visual, unique_codes):
            visuals[code] = visual
    return visuals


def check_front_detail(code: str) -> tuple[str, dict[str, object]]:
    url = public_detail_url(code)
    if not url:
        return code, {"originalUrl": "", "originalAvailable": False}
    try:
        req = Request(url, method="HEAD", headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=10) as response:
            available = 200 <= response.status < 400
        return code, {"originalUrl": url, "originalAvailable": available}
    except Exception:
        try:
            req = Request(url, headers={"User-Agent": "Mozilla/5.0", "Range": "bytes=0-256"})
            with urlopen(req, timeout=10) as response:
                available = 200 <= response.status < 400
            return code, {"originalUrl": url, "originalAvailable": available}
        except Exception:
            return code, {"originalUrl": url, "originalAvailable": False}


def check_front_details(codes: list[str]) -> dict[str, dict[str, object]]:
    unique_codes = list(dict.fromkeys(codes))
    result: dict[str, dict[str, object]] = {}
    with ThreadPoolExecutor(max_workers=18) as executor:
        for code, status in executor.map(check_front_detail, unique_codes):
            result[code] = status
    return result


def load_records() -> list[dict]:
    register_files = sorted(PLATFORM_DIR.glob("03_*.xlsx"))
    if not register_files:
        raise FileNotFoundError(f"Missing data register in {PLATFORM_DIR}")
    wb = load_workbook(register_files[0], read_only=True, data_only=True)

    clean_ws = find_sheet(wb, "سجل_منظف", 7)
    header = [str(cell or "").strip() for cell in next(clean_ws.iter_rows(values_only=True))]
    raw_rows = []
    for values in clean_ws.iter_rows(values_only=True):
        row = {header[idx]: clean_value(value) for idx, value in enumerate(values)}
        if row.get("نوع المعاملة") in CORE_TYPES:
            raw_rows.append(row)

    house_ws = find_sheet(wb, "البيوت_للبيع", 5)
    house_codes = set()
    first = True
    for values in house_ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        code = str(values[0] or "").strip()
        if code:
            house_codes.add(code)

    visible_codes = []
    records = []
    for row in raw_rows:
        code = str(row.get("الكود") or "").strip()
        if code:
            visible_codes.append(code)
        transaction = str(row.get("نوع المعاملة") or "").strip()
        price = as_float(row.get("السعر"))
        records.append(
            {
                "code": code,
                "transaction": transaction,
                "property_type": str(row.get("نوع العقار") or "").strip() or "غير محدد",
                "detail_class": str(row.get("التصنيف التفصيلي") or "").strip() or "غير محدد",
                "governorate": str(row.get("المحافظة") or "").strip() or "غير محدد",
                "area": str(row.get("المنطقة") or "").strip() or "غير محدد",
                "price": price,
                "priceText": fmt_price(price, "د.ك/شهر" if transaction in {"للإيجار", "مطلوب للإيجار"} else "د.ك"),
                "priceStatus": str(row.get("حالة السعر") or "").strip() or "غير معلن",
                "space": as_float(row.get("المساحة")),
                "listingMode": str(row.get("نوع الإعلان") or "").strip() or "غير محدد",
                "summary": str(row.get("وصف/حالة مختصرة") or "").strip(),
                "features": str(row.get("ملامح مستخرجة") or "").strip(),
                "hasImage": str(row.get("صورة") or "").strip() or "لا",
                "publishedDate": str(row.get("تاريخ النشر") or "").strip(),
                "daysSince": row.get("أيام منذ النشر") if row.get("أيام منذ النشر") != "" else "",
                "detailAvailable": str(row.get("تفاصيل متاحة") or "").strip() or "لا",
                "source": local_detail_url(code),
                "searchUrl": original_search_url(code),
                "originalUrl": public_detail_url(code),
                "originalAvailable": False,
                "imageUrl": "",
                "detailTitle": "",
                "detailText": "",
                "isHouseSale": code in house_codes,
            }
        )

    visuals = fetch_listing_visuals(visible_codes)
    front_details = check_front_details(visible_codes)
    for row in records:
        row.update(visuals.get(row["code"], {}))
        row.update(front_details.get(row["code"], {}))
        if row.get("detailTitle") and not row.get("summary"):
            row["summary"] = row["detailTitle"]
        if row.get("detailText") and not row.get("features"):
            row["features"] = row["detailText"]
        if row.get("detailTitle") or row.get("detailText"):
            row["detailAvailable"] = "نعم"
    available_records = [row for row in records if row.get("originalAvailable")]
    if available_records:
        records = available_records
    records.sort(key=lambda item: (item["publishedDate"], item["code"]), reverse=True)
    return records


def filtered(records: list[dict], metric: str, governorate: str | None = None) -> list[dict]:
    result = []
    for row in records:
        ok = (
            metric == "movement"
            or (metric == "sell" and row["transaction"] == "للبيع")
            or (metric == "buy" and row["transaction"] == "مطلوب للشراء")
            or (metric == "house" and row["isHouseSale"])
            or (metric == "rent" and row["transaction"] == "للإيجار")
            or (metric == "rent_request" and row["transaction"] == "مطلوب للإيجار")
        )
        if ok and (not governorate or row["governorate"] == governorate):
            result.append(row)
    return result


def is_known_governorate(value: object) -> bool:
    return str(value or "").strip() not in {"", "غير محدد"}


def governorate_summary(records: list[dict]) -> list[dict]:
    governorates = sorted({row["governorate"] for row in records if is_known_governorate(row.get("governorate"))})
    rows = []
    for gov in governorates:
        rows.append(
            {
                "governorate": gov,
                "movement": len(filtered(records, "movement", gov)),
                "sell": len(filtered(records, "sell", gov)),
                "buy": len(filtered(records, "buy", gov)),
                "house": len(filtered(records, "house", gov)),
                "rent": len(filtered(records, "rent", gov)),
                "rent_request": len(filtered(records, "rent_request", gov)),
            }
        )
    rows.sort(key=lambda item: item["movement"], reverse=True)
    return rows


def price_summary(rows: list[dict]) -> dict:
    values = [row["price"] for row in rows if row.get("price")]
    return {
        "priced": len(values),
        "unpriced": len(rows) - len(values),
        "avg": round(mean(values), 1) if values else None,
        "median": round(median(values), 1) if values else None,
    }


def metric_summary(records: list[dict]) -> list[dict]:
    metrics = [
        ("movement", "حركة الدلال", "إجمالي السجلات ضمن البنود الأربعة المطلوبة"),
        ("sell", "عروض للبيع", "كل عروض البيع المنشورة ضمن السجل"),
        ("buy", "طلبات شراء", "طلبات الشراء المنشورة ضمن السجل"),
        ("house", "البيوت للبيع", "العروض المصنفة ضمن البيوت للبيع"),
        ("rent", "عروض الإيجار", "عروض الإيجار المنشورة ضمن السجل"),
        ("rent_request", "طلبات الإيجار", "طلبات الإيجار المنشورة ضمن السجل"),
    ]
    rows = []
    for metric, label, note in metrics:
        items = filtered(records, metric)
        stats = price_summary(items)
        rows.append(
            {
                "metric": metric,
                "label": label,
                "count": len(items),
                "note": note,
                "priced": stats["priced"],
                "unpriced": stats["unpriced"],
                "avg": stats["avg"],
                "median": stats["median"],
            }
        )
    return rows


def write_table(ws, headers: list[str], rows: list[list], table_name: str, widths: list[int]) -> None:
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    header_fill = PatternFill("solid", fgColor=BRAND["navy"])
    header_font = Font(name="Arial", size=12, color="FFFFFF", bold=True)
    body_font = Font(name="Arial", size=11, color=BRAND["ink"])
    thin = Side(style="thin", color=BRAND["line"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            if isinstance(cell.value, str) and (cell.value.startswith("http") or cell.value.endswith(".html")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=table_name, ref=ws.dimensions)
        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)


def record_rows(rows: list[dict]) -> list[list]:
    return [
        [
            row["code"],
            row["transaction"],
            row["governorate"],
            row["area"],
            row["property_type"],
            row["detail_class"],
            row["price"] or "",
            row["priceStatus"],
            row["space"] or "",
            row["listingMode"],
            row["summary"],
            row["features"],
            row["publishedDate"],
            row["detailAvailable"],
            row.get("originalUrl", "") if row.get("originalAvailable") else "",
            "نعم" if row.get("originalAvailable") else "لا",
            row.get("originalUrl", ""),
            row.get("imageUrl", ""),
        ]
        for row in rows
    ]


def esc(value: object) -> str:
    return html_lib.escape(str(value or ""))


def create_detail_pages(records: list[dict]) -> None:
    DETAILS_DIR.mkdir(parents=True, exist_ok=True)
    for old_file in DETAILS_DIR.glob("AF-*.html"):
        old_file.unlink()
    for row in records:
        image_html = (
            f'<img src="{esc(row.get("imageUrl"))}" alt="{esc(row["area"])}" loading="lazy">'
            if row.get("imageUrl")
            else '<div class="no-image">بدون صورة</div>'
        )
        features = esc(row.get("features"))
        summary = esc(row.get("summary"))
        detail_text = esc(row.get("detailText"))
        detail_title = esc(row.get("detailTitle"))
        detail_parts = []
        if detail_text:
            detail_parts.append(f"<p>{detail_text}</p>")
        elif detail_title:
            detail_parts.append(f"<p>{detail_title}</p>")
        else:
            if summary:
                detail_parts.append(f"<p>{summary}</p>")
            if features:
                detail_parts.append(f"<p>{features}</p>")
        if not detail_parts:
            detail_parts.append('<p class="muted">لا توجد تفاصيل نصية إضافية مسجلة لهذا الإعلان.</p>')
        details_html = (
            '<div class="details-box"><b>تفاصيل الإعلان</b>'
            + "".join(detail_parts)
            + '</div>'
        )
        page = f"""<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(row["code"])} - تفاصيل الإعلان</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, Tahoma, sans-serif; background: #f6f8fb; color: #111827; font-size: 18px; line-height: 1.65; }}
    header {{ background: #0f172a; color: #fff; padding: 24px 34px; border-bottom: 5px solid #d97706; }}
    header h1 {{ margin: 0; font-size: 30px; }}
    main {{ max-width: 1120px; margin: 0 auto; padding: 24px; }}
    .card {{ background: #fff; border: 1px solid #cbd5e1; border-radius: 8px; overflow: hidden; display: grid; grid-template-columns: 420px 1fr; }}
    .image {{ min-height: 360px; background: #e2e8f0; display: flex; align-items: center; justify-content: center; color: #475569; font-weight: 900; }}
    .image img {{ width: 100%; height: 100%; object-fit: cover; display: block; }}
    .body {{ padding: 22px; }}
    .badge {{ display: inline-block; background: #eef2ff; color: #2554d9; border-radius: 999px; padding: 5px 12px; font-weight: 900; margin-bottom: 10px; }}
    .facts {{ display: grid; grid-template-columns: repeat(2, minmax(180px, 1fr)); gap: 10px; margin: 16px 0; }}
    .fact {{ background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 7px; padding: 10px; }}
    .fact b {{ display: block; color: #475569; font-size: 13px; }}
    .fact span {{ display: block; font-size: 20px; font-weight: 900; margin-top: 3px; }}
    .text {{ color: #334155; font-weight: 700; margin-top: 14px; }}
    .details-box {{ margin-top: 14px; border: 1px solid #dbe5f0; border-right: 4px solid #2554d9; border-radius: 7px; background: #f8fafc; padding: 12px 14px; }}
    .details-box b {{ display: block; color: #111827; font-size: 16px; margin-bottom: 6px; }}
    .details-box p {{ margin: 5px 0; color: #334155; font-weight: 700; white-space: pre-line; }}
    .details-box .muted {{ color: #64748b; }}
    .actions {{ display: flex; gap: 10px; flex-wrap: wrap; margin-top: 18px; }}
    a {{ display: inline-flex; align-items: center; min-height: 40px; border-radius: 6px; padding: 8px 14px; text-decoration: none; font-weight: 900; }}
    .primary {{ background: #2554d9; color: #fff; }}
    .secondary {{ display: inline-flex; align-items: center; min-height: 40px; border-radius: 6px; padding: 8px 14px; text-decoration: none; font-weight: 900; border: 1px solid #cbd5e1; color: #111827; background: #fff; }}
    @media (max-width: 820px) {{ .card {{ grid-template-columns: 1fr; }} .image {{ min-height: 240px; }} .facts {{ grid-template-columns: 1fr; }} }}
  </style>
</head>
<body>
  <header>
    <h1>{esc(row["code"])} - {esc(row["area"])}</h1>
  </header>
  <main>
    <section class="card">
      <div class="image">{image_html}</div>
      <div class="body">
        <span class="badge">{esc(row["transaction"])}</span>
        <div class="facts">
          <div class="fact"><b>كود الإعلان</b><span>{esc(row["code"])}</span></div>
          <div class="fact"><b>المحافظة</b><span>{esc(row["governorate"])}</span></div>
          <div class="fact"><b>المنطقة</b><span>{esc(row["area"])}</span></div>
          <div class="fact"><b>نوع العقار</b><span>{esc(row["property_type"])}</span></div>
          <div class="fact"><b>التصنيف</b><span>{esc(row["detail_class"])}</span></div>
          <div class="fact"><b>السعر</b><span>{esc(row["priceText"])}</span></div>
          <div class="fact"><b>المساحة</b><span>{esc(str(row["space"]) + " م²" if row.get("space") else "غير محدد")}</span></div>
          <div class="fact"><b>نوع الإعلان</b><span>{esc(row["listingMode"])}</span></div>
          <div class="fact"><b>تاريخ النشر</b><span>{esc(row["publishedDate"] or "غير محدد")}</span></div>
        </div>
        {details_html}
        <div class="actions">
          <a class="primary" href="../01_لوحة_تفاعلية_للأرقام_والعروض.html">العودة للوحة</a>
          {f'<a class="secondary" href="{esc(row.get("originalUrl"))}" target="_blank" rel="noopener">فتح صفحة الإعلان الأصلية</a>' if row.get("originalAvailable") else '<span class="secondary">صفحة الإعلان الأصلية غير متاحة</span>'}
        </div>
      </div>
    </section>
  </main>
</body>
</html>
"""
        (DETAILS_DIR / f"{row['code']}.html").write_text(page, encoding="utf-8")


def create_excel(records: list[dict], metrics: list[dict], governors: list[dict]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    detail_sheets = [
        ("كل_الحركة", "movement", "AllMovement"),
        ("عروض_بيع", "sell", "SaleOffers"),
        ("طلبات_شراء", "buy", "BuyRequests"),
        ("بيوت_للبيع", "house", "HouseSale"),
        ("عروض_إيجار", "rent", "RentOffers"),
        ("طلبات_إيجار", "rent_request", "RentRequests"),
    ]
    headers = [
        "الكود",
        "نوع المعاملة",
        "المحافظة",
        "المنطقة",
        "نوع العقار",
        "التصنيف",
        "السعر",
        "حالة السعر",
        "المساحة",
        "نوع الإعلان",
        "وصف مختصر",
        "ملامح",
        "تاريخ النشر",
        "تفاصيل متاحة",
        "صفحة الإعلان الأصلية",
        "الأصلية متاحة",
        "الموقع الأصلي",
        "رابط الصورة",
    ]
    widths = [14, 16, 22, 22, 16, 20, 14, 14, 12, 16, 38, 34, 16, 16, 58, 14, 42, 58]
    for sheet_name, metric, table_name in detail_sheets:
        ws = wb.create_sheet(sheet_name)
        write_table(ws, headers, record_rows(filtered(records, metric)), table_name, widths)

    index_rows = [
        [
            row["label"],
            row["count"],
            row["priced"],
            row["unpriced"],
            fmt_price(row["avg"]) if row["avg"] else "",
            fmt_price(row["median"]) if row["median"] else "",
            row["note"],
        ]
        for row in metrics
    ]
    ws = wb.create_sheet("فهرس_الأرقام")
    write_table(
        ws,
        ["المحور", "العدد", "أسعار معلنة", "أسعار غير معلنة", "متوسط السعر", "وسيط السعر", "ملاحظة"],
        index_rows,
        "IndexNumbers",
        [22, 12, 14, 16, 18, 18, 52],
    )

    gov_rows = [
        [row["governorate"], row["movement"], row["sell"], row["buy"], row["house"], row["rent"], row["rent_request"]]
        for row in governors
    ]
    ws = wb.create_sheet("حسب_المحافظات")
    write_table(
        ws,
        ["المحافظة", "إجمالي الحركة", "للبيع", "مطلوب شراء", "بيوت للبيع", "إيجار", "طلب إيجار"],
        gov_rows,
        "Governorates",
        [22, 16, 12, 16, 15, 12, 14],
    )

    wb.active = 0
    wb.save(XLSX_PATH)


def update_parent_register_links(records: list[dict]) -> None:
    register_files = sorted(PLATFORM_DIR.glob("03_*.xlsx"))
    if not register_files:
        return
    wb = load_workbook(register_files[0])
    detail_header = "صفحة التفاصيل"
    changed = False
    for ws in wb.worksheets:
        detail_cols = [
            col_idx
            for col_idx in range(1, ws.max_column + 1)
            if str(ws.cell(1, col_idx).value or "").strip() == detail_header
        ]
        for col_idx in reversed(detail_cols):
            ws.delete_cols(col_idx, 1)
            changed = True
    if changed:
        wb.save(register_files[0])


def remove_detail_pages() -> None:
    if DETAILS_DIR.exists():
        shutil.rmtree(DETAILS_DIR)


def create_property_type_variant(html: str) -> str:
    variant_script = """
    const propertyVariantMetrics = ['movement', 'sell', 'buy', 'rent', 'rent_request'];
    const propertyVariantHeaders = {
      movement: 'الحركة',
      sell: 'للبيع',
      buy: 'شراء',
      rent: 'إيجار',
      rent_request: 'طلب إيجار'
    };
    function propertyVariantTableRows(governorate) {
      const transaction = document.getElementById('transactionFilter').value;
      const property = document.getElementById('propertyFilter').value;
      const area = document.getElementById('areaFilter').value;
      const priceMode = document.getElementById('priceFilter').value;
      const q = normalizeText(document.getElementById('searchBox').value);
      return records.filter(row => {
        if (governorate && row.governorate !== governorate) return false;
        if (!matchesChoice(row.transaction, transaction)) return false;
        if (!matchesChoice(row.property_type, property)) return false;
        if (!matchesChoice(row.area, area)) return false;
        if (priceMode === 'priced' && !row.price) return false;
        if (priceMode === 'unpriced' && row.price) return false;
        if (!q) return true;
        return normalizeText([
          row.code, row.transaction, row.property_type, row.detail_class,
          row.governorate, row.area, row.priceText, row.listingMode,
          row.summary, row.features, row.detailText, row.publishedDate
        ].join(' ')).includes(q);
      });
    }
    renderMetrics = function() {
      const grid = document.getElementById('metricGrid');
      grid.innerHTML = '';
      payload.metrics
        .filter(item => item.metric !== 'house')
        .forEach(item => {
          const btn = document.createElement('button');
          btn.className = 'metric';
          btn.dataset.metric = item.metric;
          const avg = item.avg ? `متوسط: ${formatPrice(item.avg)}` : `أسعار معلنة: ${formatNumber(item.priced)}`;
          btn.innerHTML = `<span>${item.label}</span><strong>${formatNumber(item.count)}</strong><small>${avg}</small><em>اضغط للاختيار</em>`;
          btn.addEventListener('click', () => applyFilter(item.metric));
          grid.appendChild(btn);
        });
    };
    renderQuickList = function() {
      const select = document.getElementById('quickList');
      select.innerHTML = '';
      payload.metrics
        .filter(item => item.metric !== 'house')
        .forEach(item => {
          select.add(new Option(`${item.label} (${formatNumber(item.count)})`, `${item.metric}||`));
        });
      payload.governors.forEach(gov => {
        propertyVariantMetrics.forEach(metric => {
          const count = propertyVariantTableRows(gov.governorate).filter(row => matchesMetric(row, metric)).length;
          if (!count) return;
          select.add(new Option(`${gov.governorate} - ${metricLabels[metric]} (${formatNumber(count)})`, `${metric}||${gov.governorate}`));
        });
      });
      select.addEventListener('change', () => {
        const [metric, governorate] = select.value.split('||');
        applyFilter(metric, governorate);
      });
    };
    updateGovernorateFilterSummary = function() {
      const target = document.getElementById('govFilterSummary');
      if (!target) return;
      const pairs = [];
      const transaction = document.getElementById('transactionFilter').value;
      const property = document.getElementById('propertyFilter').value;
      const area = document.getElementById('areaFilter').value;
      const priceText = document.getElementById('priceFilter').selectedOptions[0].textContent;
      const query = document.getElementById('searchBox').value.trim();
      if (state.metric) pairs.push(['المحور', metricLabels[state.metric]]);
      if (transaction) pairs.push(['نوع المعاملة', transaction]);
      if (property) pairs.push(['نوع العقار', property]);
      if (area) pairs.push(['المنطقة', area]);
      if (priceText !== 'كل الأسعار') pairs.push(['حالة السعر', priceText]);
      if (query) pairs.push(['بحث', query]);
      target.textContent = pairs.length ? pairs.map(([label, value]) => `${label}: ${value}`).join(' | ') : 'كل أنواع العقار وكل المحافظات';
    };
    renderGovernors = function() {
      const table = document.getElementById('govTable');
      const thead = table.querySelector('thead');
      thead.innerHTML = `<tr><th>المحافظة</th>${propertyVariantMetrics.map(metric => `<th data-metric="${metric}">${propertyVariantHeaders[metric]}</th>`).join('')}</tr>`;
      const tbody = table.querySelector('tbody');
      tbody.innerHTML = '';
      const governorates = [...new Set(records.map(row => row.governorate).filter(value => value && value !== 'غير محدد'))]
        .sort((a, b) => String(a).localeCompare(String(b), 'ar'));
      governorates.forEach(governorate => {
        const rows = propertyVariantTableRows(governorate);
        const total = rows.length;
        if (!total) return;
        const tr = document.createElement('tr');
        tr.dataset.governorate = governorate;
        tr.innerHTML = `<td>${governorate}</td>`;
        propertyVariantMetrics.forEach(metric => {
          const count = rows.filter(row => matchesMetric(row, metric)).length;
          const td = document.createElement('td');
          td.dataset.metric = metric;
          const btn = document.createElement('button');
          btn.className = 'count-button';
          btn.textContent = formatNumber(count);
          btn.disabled = !count;
          btn.addEventListener('click', () => applyFilter(metric, governorate));
          td.appendChild(btn);
          tr.appendChild(td);
        });
        tbody.appendChild(tr);
      });
      highlightGovernorate();
      highlightMetricColumn();
    };
    highlightMetricColumn = function() {
      const table = document.getElementById('govTable');
      table.className = `metric-${state.metric}`;
      document.querySelectorAll('#govTable [data-metric]').forEach(cell => {
        cell.classList.toggle('metric-column-active', cell.dataset.metric === state.metric);
      });
    };
    const basePropertyVariantRenderResults = renderResults;
    renderResults = function() {
      basePropertyVariantRenderResults();
      renderGovernors();
    };
    """
    variant = html.replace(
        "<title>استعراض الأرقام والعروض الفعلية</title>",
        "<title>استعراض الأرقام والعروض الفعلية - نسخة نوع العقار</title>",
    )
    variant = variant.replace(
        "<h1>استعراض الأرقام والعروض الفعلية</h1>",
        "<h1>استعراض الأرقام والعروض الفعلية - نسخة نوع العقار</h1>",
    )
    variant = variant.replace(
        "يمكن اختيار رقم من الجدول مباشرة لعرض المحافظة والمحور المطلوب.",
        "اختياراتك من الفلاتر أعلى الصفحة ستظهر هنا، ثم اختر الرقم من جدول المحافظات لعرض السجلات الفعلية.",
    )
    variant = variant.replace('\n                <th data-metric="house">بيوت للبيع</th>', "")
    variant = variant.replace("    renderMetrics();", variant_script + "\n    renderMetrics();")
    return variant


def copy_html_assets() -> None:
    assets_dir = OUTPUT_DIR / "assets"
    assets_dir.mkdir(parents=True, exist_ok=True)
    for source in (LOGO_PATH, COVER_PATH):
        if source.exists():
            shutil.copy2(source, assets_dir / source.name)


def create_html(records: list[dict], metrics: list[dict], governors: list[dict]) -> None:
    copy_html_assets()
    payload = {
        "records": records,
        "metrics": metrics,
        "governors": governors,
        "generatedLabel": arabic_date_label(date.today()),
        "logoDataUri": "assets/alforaij_logo.png",
        "coverDataUri": "assets/kuwait_glass_cover.png",
    }
    data_json = json.dumps(payload, ensure_ascii=False)
    html = f"""<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>استعراض الأرقام والعروض الفعلية</title>
  <style>
    :root {{
      --ink: #{BRAND["ink"]};
      --muted: #{BRAND["muted"]};
      --line: #{BRAND["line"]};
      --bg: #{BRAND["bg"]};
      --panel: #{BRAND["panel"]};
      --navy: #{BRAND["navy"]};
      --blue: #{BRAND["blue"]};
      --orange: #{BRAND["orange"]};
      --green: #{BRAND["green"]};
      --red: #{BRAND["red"]};
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Tahoma, "Segoe UI", sans-serif;
      background: var(--bg);
      color: var(--ink);
      font-size: 18px;
      line-height: 1.55;
    }}
    header {{
      position: relative;
      min-height: 245px;
      color: #fff;
      padding: 32px 42px;
      border-bottom: 5px solid var(--orange);
      overflow: hidden;
      background:
        linear-gradient(90deg, rgba(15, 23, 42, 0.96) 0%, rgba(15, 23, 42, 0.88) 46%, rgba(15, 23, 42, 0.42) 100%),
        var(--navy);
    }}
    header::before {{
      content: "";
      position: absolute;
      inset: 0;
      background-image: var(--cover-image);
      background-size: cover;
      background-position: center;
      opacity: 0.36;
      transform: scale(1.02);
    }}
    .hero-content {{
      position: relative;
      z-index: 1;
      max-width: 1180px;
      margin: 0 auto;
      display: grid;
      grid-template-columns: minmax(0, 1.16fr) minmax(280px, 0.84fr);
      grid-template-areas: "title logo";
      gap: 34px;
      align-items: center;
      min-height: 178px;
      direction: ltr;
    }}
    .logo-showcase {{
      grid-area: logo;
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 168px;
      border: 1px solid rgba(255, 255, 255, 0.24);
      border-radius: 18px;
      background:
        linear-gradient(145deg, rgba(255, 255, 255, 0.22), rgba(255, 255, 255, 0.08)),
        rgba(15, 23, 42, 0.18);
      box-shadow:
        inset 0 1px 0 rgba(255, 255, 255, 0.28),
        0 24px 58px rgba(2, 6, 23, 0.32);
      backdrop-filter: blur(12px);
      -webkit-backdrop-filter: blur(12px);
      position: relative;
      overflow: hidden;
    }}
    .logo-showcase::before {{
      content: "";
      position: absolute;
      width: 150px;
      height: 150px;
      border-radius: 50%;
      background: rgba(255, 255, 255, 0.16);
      filter: blur(24px);
      top: -48px;
      right: -34px;
    }}
    .logo-showcase::after {{
      content: "";
      position: absolute;
      inset: 14px;
      border-radius: 14px;
      border: 1px solid rgba(255, 255, 255, 0.12);
      pointer-events: none;
    }}
    .brand-logo {{
      width: min(78%, 292px);
      min-height: 92px;
      object-fit: contain;
      background: rgba(255, 255, 255, 0.92);
      border: 1px solid rgba(255, 255, 255, 0.86);
      border-radius: 12px;
      padding: 16px 18px;
      box-shadow:
        0 18px 40px rgba(2, 6, 23, 0.28),
        inset 0 1px 0 rgba(255, 255, 255, 0.78);
      position: relative;
      z-index: 1;
    }}
    .hero-title {{
      grid-area: title;
      max-width: 930px;
      margin-left: auto;
      text-align: right;
      direction: rtl;
      background: linear-gradient(90deg, rgba(15, 23, 42, 0.02), rgba(15, 23, 42, 0.54));
      border-right: 5px solid var(--orange);
      border-radius: 8px;
      padding: 18px 24px 20px;
    }}
    .hero-title h1 {{ margin: 0; font-size: 40px; line-height: 1.22; max-width: 900px; }}
    .hero-title p {{ margin: 12px 0 0; color: #e2e8f0; font-weight: 800; max-width: 850px; font-size: 21px; }}
    main {{ max-width: 1360px; margin: 0 auto; padding: 28px; }}
    .toolbar, .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 1px 2px rgba(15, 23, 42, 0.08);
    }}
    .action-toolbar {{
      display: flex;
      flex-wrap: wrap;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 1px 2px rgba(15, 23, 42, 0.08);
      padding: 14px 16px;
      margin-bottom: 18px;
    }}
    .action-toolbar strong {{
      color: var(--ink);
      font-size: 18px;
      font-weight: 900;
    }}
    .action-toolbar span {{
      display: block;
      margin-top: 2px;
      color: var(--muted);
      font-size: 14px;
      font-weight: 800;
    }}
    .action-buttons {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }}
    .action-button {{
      min-height: 44px;
      border: 1px solid var(--line);
      border-radius: 7px;
      padding: 10px 16px;
      background: #fff;
      color: var(--ink);
      font: inherit;
      font-size: 15px;
      font-weight: 900;
      text-decoration: none;
      cursor: pointer;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
    }}
    .action-button.primary {{
      background: var(--blue);
      border-color: var(--blue);
      color: #fff;
      box-shadow: 0 8px 18px rgba(37, 84, 217, 0.18);
    }}
    .action-button:hover {{
      border-color: var(--blue);
      color: var(--blue);
    }}
    .action-button.primary:hover {{
      color: #fff;
      background: #1d4ed8;
    }}
    .toolbar {{
      display: grid;
      grid-template-columns: 1.4fr repeat(5, minmax(150px, 1fr));
      gap: 14px;
      padding: 16px;
      margin-bottom: 18px;
    }}
    .search-toolbar {{
      grid-template-columns: minmax(320px, 1.6fr) 220px 180px;
      align-items: end;
    }}
    label {{ display: block; color: var(--muted); font-size: 14px; font-weight: 800; margin-bottom: 6px; }}
    input, select {{
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 11px 12px;
      font: inherit;
      font-size: 17px;
      color: var(--ink);
      background: #fff;
    }}
    .autocomplete {{ position: relative; }}
    .suggestions {{
      display: none;
      position: absolute;
      top: calc(100% + 5px);
      left: 0;
      right: 0;
      z-index: 40;
      max-height: 245px;
      overflow: auto;
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 14px 30px rgba(15, 23, 42, 0.16);
      padding: 6px;
    }}
    .suggestions.show {{ display: block; }}
    .suggestion-item {{
      width: 100%;
      display: block;
      border: 0;
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      padding: 10px 12px;
      text-align: right;
      font: inherit;
      font-size: 16px;
      font-weight: 800;
      cursor: pointer;
    }}
    .suggestion-item:hover,
    .suggestion-item:focus {{
      background: #eef6ff;
      color: var(--blue);
      outline: none;
    }}
    .metric-grid {{
      display: grid;
      grid-template-columns: repeat(6, minmax(140px, 1fr));
      gap: 12px;
      margin-bottom: 18px;
    }}
    .metric {{
      text-align: right;
      border: 1px solid var(--line);
      background: #fff;
      border-radius: 8px;
      padding: 14px;
      cursor: pointer;
      min-height: 118px;
      border-right: 6px solid var(--blue);
    }}
    .metric[data-metric="sell"], .metric[data-metric="house"] {{ border-right-color: var(--orange); }}
    .metric[data-metric="rent"], .metric[data-metric="rent_request"] {{ border-right-color: var(--green); }}
    .metric.active {{ outline: 3px solid rgba(37, 84, 217, 0.18); border-color: var(--blue); }}
    .metric span {{ display: block; color: var(--muted); font-size: 15px; font-weight: 800; }}
    .metric strong {{ display: block; font-size: 34px; margin-top: 6px; line-height: 1; }}
    .metric small {{ display: block; margin-top: 10px; color: var(--muted); font-weight: 700; }}
    .metric em {{
      display: inline-flex;
      margin-top: 9px;
      border-radius: 999px;
      background: #eff6ff;
      color: var(--blue);
      padding: 4px 9px;
      font-style: normal;
      font-size: 13px;
      font-weight: 900;
    }}
    .metric[data-metric="sell"] em,
    .metric[data-metric="house"] em {{
      background: #fff7ed;
      color: #9a3412;
    }}
    .metric[data-metric="rent"] em,
    .metric[data-metric="rent_request"] em {{
      background: #ecfdf5;
      color: #065f46;
    }}
    .layout {{
      display: grid;
      grid-template-columns: 0.95fr 1.35fr;
      gap: 18px;
      align-items: start;
    }}
    .panel {{ padding: 18px; }}
    .panel h2 {{ margin: 0 0 14px; font-size: 24px; }}
    .panel-title-row {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
      flex-wrap: wrap;
      margin-bottom: 8px;
    }}
    .panel-title-row h2 {{ margin: 0; }}
    .filter-summary {{
      display: inline-flex;
      align-items: center;
      min-height: 30px;
      border-radius: 999px;
      border: 1px solid #bfdbfe;
      background: #eff6ff;
      color: var(--blue);
      padding: 4px 12px;
      font-size: 14px;
      font-weight: 900;
      max-width: 100%;
    }}
    .panel-hint {{
      margin: -6px 0 14px;
      color: var(--muted);
      font-size: 15px;
      font-weight: 800;
    }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 10px; text-align: center; }}
    th {{ background: var(--navy); color: #fff; font-size: 15px; }}
    td:first-child, th:first-child {{ text-align: right; }}
    .count-button {{
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      min-width: 54px;
      padding: 8px 10px;
      font-weight: 900;
      font-size: 17px;
      cursor: pointer;
    }}
    .count-button:hover {{ border-color: var(--blue); color: var(--blue); }}
    #govTable tbody tr.gov-active td {{
      background: #eef6ff;
      border-bottom-color: #bfdbfe;
      color: #0f172a;
      font-weight: 900;
    }}
    #govTable tbody tr.gov-active td:first-child {{
      border-right: 5px solid var(--blue);
    }}
    #govTable tbody tr.gov-active .count-button {{
      border-color: var(--blue);
      background: #eff6ff;
      color: var(--blue);
    }}
    #govTable .metric-column-active {{
      font-weight: 900;
      box-shadow: inset 0 0 0 999px rgba(37, 84, 217, 0.06);
    }}
    #govTable.metric-sell .metric-column-active,
    #govTable.metric-house .metric-column-active {{
      box-shadow: inset 0 0 0 999px rgba(217, 119, 6, 0.10);
    }}
    #govTable.metric-rent .metric-column-active,
    #govTable.metric-rent_request .metric-column-active {{
      box-shadow: inset 0 0 0 999px rgba(4, 120, 87, 0.10);
    }}
    #govTable th.metric-column-active {{
      position: relative;
      color: #fff;
    }}
    #govTable th.metric-column-active::after {{
      content: "";
      position: absolute;
      inset-inline-start: 10px;
      inset-inline-end: 10px;
      bottom: 4px;
      height: 3px;
      border-radius: 999px;
      background: #fbbf24;
    }}
    #govTable td.metric-column-active .count-button {{
      border-color: var(--orange);
      background: #fff7ed;
      color: #9a3412;
    }}
    #govTable.metric-rent td.metric-column-active .count-button,
    #govTable.metric-rent_request td.metric-column-active .count-button {{
      border-color: var(--green);
      background: #ecfdf5;
      color: #065f46;
    }}
    #govTable.metric-movement td.metric-column-active .count-button,
    #govTable.metric-buy td.metric-column-active .count-button {{
      border-color: var(--blue);
      background: #eff6ff;
      color: var(--blue);
    }}
    .result-head {{
      display: flex;
      gap: 12px;
      justify-content: space-between;
      align-items: flex-start;
      margin-bottom: 12px;
    }}
    .result-head strong {{ font-size: 26px; }}
    .result-tools {{
      display: grid;
      grid-template-columns: 1fr 0.6fr;
      gap: 10px;
      margin-bottom: 12px;
    }}
    .summary-strip {{
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 8px;
      margin-bottom: 12px;
    }}
    .selection-chips {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 4px 0 12px;
    }}
    .selection-chips span {{
      display: inline-flex;
      align-items: center;
      min-height: 32px;
      background: #fff7ed;
      border: 1px solid #fed7aa;
      color: #9a3412;
      border-radius: 999px;
      padding: 5px 12px;
      font-size: 14px;
      font-weight: 900;
    }}
    .summary-strip div {{
      background: #f8fafc;
      border: 1px solid #e2e8f0;
      border-radius: 7px;
      padding: 9px 10px;
    }}
    .summary-strip b {{ display: block; color: var(--muted); font-size: 12px; }}
    .summary-strip span {{ display: block; font-size: 18px; font-weight: 900; margin-top: 2px; }}
    .list {{
      display: grid;
      gap: 10px;
      max-height: 680px;
      overflow: auto;
      padding-left: 4px;
    }}
    .item {{
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fff;
      padding: 0;
      display: grid;
      grid-template-columns: 148px 1fr;
      overflow: hidden;
    }}
    .thumb {{
      min-height: 168px;
      background: linear-gradient(135deg, #e2e8f0, #f8fafc);
      display: flex;
      align-items: center;
      justify-content: center;
      color: var(--muted);
      font-size: 14px;
      font-weight: 900;
      text-align: center;
    }}
    .thumb img {{ width: 100%; height: 100%; object-fit: cover; display: block; }}
    .item-body {{ padding: 12px 14px; min-width: 0; }}
    .item-top {{
      display: flex;
      gap: 10px;
      justify-content: space-between;
      align-items: flex-start;
      border-bottom: 1px solid #e7edf5;
      padding-bottom: 8px;
      margin-bottom: 8px;
    }}
    .item h3 {{ margin: 0; font-size: 19px; }}
    .badge {{
      display: inline-flex;
      align-items: center;
      min-height: 28px;
      border-radius: 999px;
      padding: 4px 10px;
      background: #eef2ff;
      color: var(--blue);
      font-weight: 900;
      font-size: 14px;
      white-space: nowrap;
    }}
    .facts {{
      display: grid;
      grid-template-columns: repeat(4, minmax(120px, 1fr));
      gap: 8px;
      margin: 8px 0;
    }}
    .fact {{ background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px; padding: 8px; }}
    .fact b {{ display: block; color: var(--muted); font-size: 13px; }}
    .fact span {{ display: block; font-weight: 900; margin-top: 3px; }}
    .details-box {{
      margin-top: 10px;
      border: 1px solid #dbe5f0;
      border-right: 4px solid var(--blue);
      border-radius: 7px;
      background: #f8fafc;
      padding: 10px 12px;
    }}
    .details-box b {{
      display: block;
      color: var(--ink);
      font-size: 15px;
      margin-bottom: 5px;
    }}
    .details-box p {{
      margin: 4px 0;
      color: #334155;
      font-size: 15px;
      font-weight: 700;
      line-height: 1.65;
      white-space: pre-line;
    }}
    .details-box .muted {{ color: var(--muted); }}
    .desc {{ color: var(--muted); font-size: 15px; font-weight: 700; margin: 8px 0 0; }}
    .source {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 36px;
      margin-top: 9px;
      border-radius: 6px;
      background: var(--blue);
      border: 0;
      color: var(--blue);
      color: #fff;
      padding: 7px 12px;
      font-weight: 900;
      text-decoration: none;
      cursor: pointer;
    }}
    .source:hover {{ filter: brightness(0.95); }}
    .clear-button {{
      width: 100%;
      min-height: 48px;
      border: 1px solid #fecaca;
      border-radius: 6px;
      background: #fff;
      color: #991b1b;
      font: inherit;
      font-size: 17px;
      font-weight: 900;
      cursor: pointer;
    }}
    .clear-button:hover {{ background: #fef2f2; }}
    .secondary-source {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 36px;
      margin: 9px 8px 0 0;
      border-radius: 6px;
      border: 1px solid var(--line);
      background: #fff;
      color: var(--ink);
      padding: 7px 12px;
      font-weight: 900;
      text-decoration: none;
    }}
    .modal-backdrop {{
      position: fixed;
      inset: 0;
      background: rgba(15, 23, 42, 0.62);
      display: none;
      align-items: center;
      justify-content: center;
      padding: 22px;
      z-index: 50;
    }}
    .modal-backdrop.show {{ display: flex; }}
    .modal {{
      width: min(980px, 100%);
      max-height: 92vh;
      overflow: auto;
      background: #fff;
      border-radius: 8px;
      border: 1px solid var(--line);
      box-shadow: 0 24px 70px rgba(15, 23, 42, 0.28);
    }}
    .modal-head {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      padding: 14px 18px;
      border-bottom: 1px solid var(--line);
      background: #f8fafc;
    }}
    .modal-head h2 {{ margin: 0; font-size: 22px; }}
    .modal-close {{
      border: 1px solid var(--line);
      background: #fff;
      border-radius: 6px;
      width: 38px;
      height: 38px;
      font-size: 22px;
      cursor: pointer;
    }}
    .modal-body {{
      display: grid;
      grid-template-columns: 360px 1fr;
      gap: 0;
    }}
    .modal-image {{
      min-height: 360px;
      background: #e2e8f0;
      display: flex;
      align-items: center;
      justify-content: center;
      color: var(--muted);
      font-weight: 900;
      text-align: center;
    }}
    .modal-image img {{ width: 100%; height: 100%; object-fit: cover; display: block; }}
    .modal-info {{ padding: 18px; }}
    .modal-actions {{ display: flex; flex-wrap: wrap; gap: 8px; margin-top: 14px; }}
    .empty {{ padding: 22px; color: var(--muted); font-weight: 900; text-align: center; }}
    @media (max-width: 1050px) {{
      .hero-content {{
        grid-template-columns: 1fr;
        grid-template-areas: "logo" "title";
        gap: 18px;
      }}
      .logo-showcase {{ min-height: 138px; }}
      .brand-logo {{ width: min(72%, 240px); min-height: 78px; }}
      .metric-grid {{ grid-template-columns: repeat(3, 1fr); }}
      .layout {{ grid-template-columns: 1fr; }}
      .toolbar {{ grid-template-columns: 1fr 1fr; }}
      .search-toolbar {{ grid-template-columns: 1fr; }}
      .facts {{ grid-template-columns: repeat(2, 1fr); }}
      .summary-strip {{ grid-template-columns: repeat(2, 1fr); }}
    }}
    @media (max-width: 640px) {{
      body {{ font-size: 16px; }}
      header {{ padding: 16px 12px; min-height: 250px; }}
      .logo-showcase {{ min-height: 118px; border-radius: 14px; }}
      .brand-logo {{ width: min(82%, 204px); min-height: 68px; padding: 12px 14px; }}
      .hero-title {{ padding: 16px; }}
      .hero-title h1 {{ font-size: 28px; }}
      .hero-title p {{ font-size: 17px; }}
      main {{ padding: 10px; }}
      .action-toolbar {{ align-items: stretch; padding: 12px; }}
      .action-buttons {{ width: 100%; }}
      .action-button {{ flex: 1 1 100%; }}
      input, select {{ font-size: 16px; }}
      .metric-grid {{ grid-template-columns: 1fr 1fr; }}
      .metric strong {{ font-size: 30px; }}
      .facts {{ grid-template-columns: 1fr; }}
      .toolbar, .result-tools, .summary-strip {{ grid-template-columns: 1fr; }}
      .item {{ grid-template-columns: 1fr; }}
      .thumb {{ min-height: 180px; }}
      .modal-backdrop {{ align-items: stretch; padding: 8px; }}
      .modal {{ width: 100%; max-height: 96vh; border-radius: 8px; }}
      .modal-head {{ position: sticky; top: 0; z-index: 2; background: #fff; }}
      .modal-head h2 {{ font-size: 19px; }}
      .modal-body {{ grid-template-columns: 1fr; }}
      .modal-image {{ min-height: 210px; max-height: 36vh; }}
      .modal-info {{ padding: 12px; }}
      th, td {{ padding: 8px 6px; font-size: 14px; }}
    }}
    @media print {{
      body {{ background: #fff; }}
      header {{ min-height: 0; padding: 20px 24px; }}
      main {{ padding: 18px; max-width: none; }}
      .action-toolbar, .toolbar, .modal, .modal-backdrop, .item-actions, .clear-button {{ display: none !important; }}
      .layout {{ grid-template-columns: 1fr; }}
      .panel, .metric, .item {{ box-shadow: none; break-inside: avoid; }}
      .list {{ max-height: none; overflow: visible; }}
    }}
  </style>
</head>
<body>
  <header style="--cover-image: url('{payload["coverDataUri"]}')">
    <div class="hero-content">
      <div class="hero-title">
        <h1>لوحة الأرقام والعروض الفعلية</h1>
        <p>استعراض تفاعلي لحركة الدلال، عروض البيع والشراء، البيوت المعروضة، والإيجارات حسب المحافظة والمنطقة ونوع العقار.</p>
      </div>
      <div class="logo-showcase" aria-label="شعار شركة عبدالعزيز سعود الفريج العقارية">
        <img class="brand-logo" src="{payload["logoDataUri"]}" alt="شركة عبدالعزيز سعود الفريج العقارية">
      </div>
    </div>
  </header>
  <main>
    <section class="action-toolbar" aria-label="أدوات التصدير">
      <div>
        <strong>ملف تفاعلي واحد للأرقام والعروض</strong>
        <span>يمكن استعراض البيانات من الصفحة، أو تحميل الجدول، أو حفظ العرض الحالي كملف PDF.</span>
      </div>
      <div class="action-buttons">
        <a class="action-button primary" href="downloads/offer-evidence.xlsx" download>تحميل Excel التفصيلي</a>
        <button id="printPdfButton" class="action-button" type="button">طباعة / حفظ PDF</button>
      </div>
    </section>

    <section class="toolbar" aria-label="أدوات التصفية">
      <div>
        <label for="quickList">قائمة جاهزة حسب الرقم</label>
        <select id="quickList"></select>
      </div>
      <div>
        <label for="governorateFilter">المحافظة - اكتب أو اختر</label>
        <div class="autocomplete">
          <input id="governorateFilter" placeholder="كل المحافظات" autocomplete="off">
          <div id="governorateFilterSuggest" class="suggestions" role="listbox"></div>
        </div>
      </div>
      <div>
        <label for="transactionFilter">نوع المعاملة - اكتب أو اختر</label>
        <div class="autocomplete">
          <input id="transactionFilter" placeholder="كل المعاملات" autocomplete="off">
          <div id="transactionFilterSuggest" class="suggestions" role="listbox"></div>
        </div>
      </div>
      <div>
        <label for="propertyFilter">نوع العقار - اكتب أو اختر</label>
        <div class="autocomplete">
          <input id="propertyFilter" placeholder="كل أنواع العقار" autocomplete="off">
          <div id="propertyFilterSuggest" class="suggestions" role="listbox"></div>
        </div>
      </div>
      <div>
        <label for="areaFilter">المنطقة - اكتب أو اختر</label>
        <div class="autocomplete">
          <input id="areaFilter" placeholder="كل المناطق" autocomplete="off">
          <div id="areaFilterSuggest" class="suggestions" role="listbox"></div>
        </div>
      </div>
      <div>
        <label for="priceFilter">حالة السعر</label>
        <select id="priceFilter">
          <option value="all">كل الأسعار</option>
          <option value="priced">أسعار معلنة فقط</option>
          <option value="unpriced">غير معلن فقط</option>
        </select>
      </div>
    </section>

    <section class="toolbar search-toolbar" aria-label="البحث والترتيب">
      <div>
        <label for="searchBox">ابحث هنا داخل النتائج</label>
        <input id="searchBox" type="search" placeholder="مثال: المطلاع، بيت، أبو فطيرة، AF-303، 400">
      </div>
      <div>
        <label for="sortFilter">الترتيب</label>
        <select id="sortFilter">
          <option value="newest">الأحدث أولا</option>
          <option value="price_desc">السعر الأعلى</option>
          <option value="price_asc">السعر الأقل</option>
          <option value="area">المنطقة أبجديا</option>
        </select>
      </div>
      <div>
        <label>&nbsp;</label>
        <button id="clearFilters" class="clear-button" type="button">مسح الاختيارات</button>
      </div>
    </section>

    <section id="metricGrid" class="metric-grid" aria-label="الأرقام الرئيسية"></section>

    <section class="layout">
      <div class="panel">
        <div class="panel-title-row">
          <h2>الأرقام حسب المحافظات</h2>
          <span id="govFilterSummary" class="filter-summary">كل السجلات</span>
        </div>
        <p class="panel-hint">يمكن اختيار رقم من الجدول مباشرة لعرض المحافظة والمحور المطلوب.</p>
        <div style="overflow:auto">
          <table id="govTable">
            <thead>
              <tr>
                <th>المحافظة</th>
                <th data-metric="movement">الحركة</th>
                <th data-metric="sell">للبيع</th>
                <th data-metric="buy">شراء</th>
                <th data-metric="house">بيوت للبيع</th>
                <th data-metric="rent">إيجار</th>
                <th data-metric="rent_request">طلب إيجار</th>
              </tr>
            </thead>
            <tbody></tbody>
          </table>
        </div>
      </div>
      <div class="panel">
        <div class="result-head">
          <div>
            <h2 id="resultTitle">كل الحركة</h2>
            <span id="resultMeta"></span>
          </div>
          <strong id="resultCount">0</strong>
        </div>
        <div id="selectionChips" class="selection-chips"></div>
        <div id="resultStats" class="summary-strip"></div>
        <div id="resultList" class="list"></div>
      </div>
    </section>
  </main>

  <div id="detailsModal" class="modal-backdrop" role="dialog" aria-modal="true" aria-labelledby="modalTitle">
    <div class="modal">
      <div class="modal-head">
        <h2 id="modalTitle">تفاصيل الإعلان</h2>
        <button id="modalClose" class="modal-close" type="button" aria-label="إغلاق">×</button>
      </div>
      <div id="modalBody" class="modal-body"></div>
    </div>
  </div>

  <script id="payload" type="application/json">{data_json}</script>
  <script>
    const payload = JSON.parse(document.getElementById('payload').textContent);
    const records = payload.records;
    const metricLabels = {{
      movement: 'حركة الدلال',
      sell: 'عروض للبيع',
      buy: 'طلبات شراء',
      house: 'البيوت للبيع',
      rent: 'عروض الإيجار',
      rent_request: 'طلبات الإيجار'
    }};
    const areaGovernorateMap = records.reduce((map, row) => {{
      if (row.area && row.governorate && !map[row.area]) map[row.area] = row.governorate;
      return map;
    }}, {{}});
    const searchableFilterIds = ['governorateFilter', 'transactionFilter', 'propertyFilter', 'areaFilter'];
    const choiceOptions = {{}};
    let state = {{ metric: 'movement', governorate: '', label: 'حركة الدلال' }};

    function formatNumber(value) {{
      return new Intl.NumberFormat('ar-KW').format(value || 0);
    }}
    function formatPrice(value) {{
      if (!value) return 'غير معلن';
      if (value >= 10000) return `${{formatNumber(Math.round(value / 1000))}} ألف د.ك`;
      return `${{formatNumber(Math.round(value))}} د.ك`;
    }}
    function uniqueValues(field) {{
      return [...new Set(records.map(row => row[field]).filter(Boolean))].sort((a, b) => String(a).localeCompare(String(b), 'ar'));
    }}
    function populateSelect(id, values, allLabel) {{
      const input = document.getElementById(id);
      if (!input) return;
      choiceOptions[id] = values;
      input.placeholder = allLabel;
      renderSuggestions(id);
    }}
    function normalizeText(value) {{
      return String(value || '')
        .trim()
        .toLowerCase()
        .replace(/[إأآ]/g, 'ا')
        .replace(/ى/g, 'ي')
        .replace(/ـ/g, '')
        .replace(/[ًٌٍَُِّْ]/g, '');
    }}
    function matchesChoice(rowValue, typedValue) {{
      const typed = normalizeText(typedValue);
      if (!typed) return true;
      const value = normalizeText(rowValue);
      return value === typed || value.includes(typed);
    }}
    function matchingChoices(id) {{
      const input = document.getElementById(id);
      const typed = normalizeText(input?.value);
      const values = choiceOptions[id] || [];
      if (!typed) return values.slice(0, 12);
      return values
        .filter(value => normalizeText(value).includes(typed))
        .slice(0, 18);
    }}
    function hideSuggestions(id) {{
      const box = document.getElementById(`${{id}}Suggest`);
      if (box) box.classList.remove('show');
    }}
    function hideAllSuggestions(exceptId = '') {{
      searchableFilterIds
        .filter(id => id !== exceptId)
        .forEach(id => hideSuggestions(id));
    }}
    function chooseSuggestion(id, value) {{
      const input = document.getElementById(id);
      if (!input) return;
      input.value = value;
      hideSuggestions(id);
      handleFilterChange(id);
    }}
    function renderSuggestions(id) {{
      if (!searchableFilterIds.includes(id)) return;
      const input = document.getElementById(id);
      const box = document.getElementById(`${{id}}Suggest`);
      if (!input || !box) return;
      hideAllSuggestions(id);
      box.innerHTML = '';
      const values = matchingChoices(id);
      if (document.activeElement !== input || !values.length) {{
        box.classList.remove('show');
        return;
      }}
      values.forEach(value => {{
        const option = document.createElement('button');
        option.type = 'button';
        option.className = 'suggestion-item';
        option.setAttribute('role', 'option');
        option.textContent = value;
        option.addEventListener('mousedown', event => {{
          event.preventDefault();
          chooseSuggestion(id, value);
        }});
        box.appendChild(option);
      }});
      box.classList.add('show');
    }}
    function updateAreaOptions(preserve = true) {{
      const areaSelect = document.getElementById('areaFilter');
      const currentArea = preserve ? areaSelect.value : '';
      const gov = document.getElementById('governorateFilter').value || state.governorate;
      const areas = [...new Set(records
        .filter(row => matchesChoice(row.governorate, gov))
        .map(row => row.area)
        .filter(Boolean))]
        .sort((a, b) => String(a).localeCompare(String(b), 'ar'));
      populateSelect('areaFilter', areas, 'كل المناطق');
      if (!preserve) {{
        areaSelect.value = '';
        renderSuggestions('areaFilter');
        return;
      }}
      if (!currentArea) return;
      const exactArea = areas.find(area => normalizeText(area) === normalizeText(currentArea));
      if (exactArea) {{
        areaSelect.value = exactArea;
        return;
      }}
      if (document.activeElement === areaSelect && areas.some(area => matchesChoice(area, currentArea))) {{
        areaSelect.value = currentArea;
        return;
      }}
      areaSelect.value = '';
    }}
    function syncGovernorateFromArea() {{
      const area = document.getElementById('areaFilter').value;
      if (!area) return;
      let governorate = areaGovernorateMap[area];
      if (!governorate) {{
        const typed = normalizeText(area);
        const matchingAreas = Object.keys(areaGovernorateMap).filter(name => normalizeText(name).includes(typed));
        if (matchingAreas.length !== 1) return;
        governorate = areaGovernorateMap[matchingAreas[0]];
      }}
      state.governorate = '';
      document.getElementById('governorateFilter').value = governorate;
    }}
    function metricFromTransactionValue(value) {{
      const normalized = normalizeText(value);
      if (!normalized) return '';
      const exactChoice = (choiceOptions.transactionFilter || []).find(option => normalizeText(option) === normalized);
      const text = exactChoice ? normalizeText(exactChoice) : normalized;
      if (text.includes('مطلوب') && text.includes('ايجار')) return 'rent_request';
      if (text.includes('مطلوب') && text.includes('شراء')) return 'buy';
      if (!text.includes('مطلوب') && text.includes('للايجار')) return 'rent';
      if (!text.includes('مطلوب') && text.includes('للبيع')) return 'sell';
      return '';
    }}
    function syncMetricFromTransaction() {{
      const metric = metricFromTransactionValue(document.getElementById('transactionFilter').value);
      if (!metric) return;
      state.metric = metric;
      const governorate = selectedGovernorate();
      state.label = governorate ? `${{governorate}} - ${{metricLabels[metric]}}` : metricLabels[metric];
      document.querySelectorAll('.metric').forEach(btn => {{
        btn.classList.toggle('active', btn.dataset.metric === metric);
      }});
    }}
    function transactionForMetric(metric) {{
      if (metric === 'sell') return 'للبيع';
      if (metric === 'buy') return 'مطلوب للشراء';
      if (metric === 'rent') return 'للإيجار';
      if (metric === 'rent_request') return 'مطلوب للإيجار';
      return '';
    }}
    function syncTransactionFromMetric(metric) {{
      const input = document.getElementById('transactionFilter');
      if (!input) return;
      input.value = transactionForMetric(metric);
    }}
    function selectedGovernorate() {{
      return state.governorate || document.getElementById('governorateFilter').value || '';
    }}
    function matchesMetric(row, metric) {{
      if (metric === 'movement') return true;
      if (metric === 'sell') return row.transaction === 'للبيع';
      if (metric === 'buy') return row.transaction === 'مطلوب للشراء';
      if (metric === 'house') return row.isHouseSale;
      if (metric === 'rent') return row.transaction === 'للإيجار';
      if (metric === 'rent_request') return row.transaction === 'مطلوب للإيجار';
      return true;
    }}
    function currentRows() {{
      const q = document.getElementById('searchBox').value.trim().toLowerCase();
      const priceMode = document.getElementById('priceFilter').value;
      const transaction = document.getElementById('transactionFilter').value;
      const governorate = selectedGovernorate();
      const property = document.getElementById('propertyFilter').value;
      const area = document.getElementById('areaFilter').value;
      const sortMode = document.getElementById('sortFilter').value;
      const rows = records.filter(row => {{
        if (!matchesMetric(row, state.metric)) return false;
        if (!matchesChoice(row.governorate, governorate)) return false;
        if (!matchesChoice(row.transaction, transaction)) return false;
        if (!matchesChoice(row.property_type, property)) return false;
        if (!matchesChoice(row.area, area)) return false;
        if (priceMode === 'priced' && !row.price) return false;
        if (priceMode === 'unpriced' && row.price) return false;
        if (!q) return true;
        return [
          row.code, row.transaction, row.property_type, row.detail_class,
          row.governorate, row.area, row.priceText, row.listingMode,
          row.summary, row.features, row.publishedDate
        ].join(' ').toLowerCase().includes(q);
      }});
      rows.sort((a, b) => {{
        if (sortMode === 'price_desc') return (b.price || -1) - (a.price || -1);
        if (sortMode === 'price_asc') return (a.price || Number.MAX_SAFE_INTEGER) - (b.price || Number.MAX_SAFE_INTEGER);
        if (sortMode === 'area') return String(a.area).localeCompare(String(b.area), 'ar') || String(b.publishedDate).localeCompare(String(a.publishedDate));
        return String(b.publishedDate).localeCompare(String(a.publishedDate)) || String(b.code).localeCompare(String(a.code));
      }});
      return rows;
    }}
    function applyFilter(metric, governorate = '') {{
      state = {{
        metric,
        governorate,
        label: governorate ? `${{governorate}} - ${{metricLabels[metric]}}` : metricLabels[metric]
      }};
      document.getElementById('governorateFilter').value = governorate || '';
      syncTransactionFromMetric(metric);
      updateAreaOptions(false);
      document.querySelectorAll('.metric').forEach(btn => {{
        btn.classList.toggle('active', btn.dataset.metric === metric);
      }});
      highlightGovernorate();
      highlightMetricColumn();
      updateGovernorateFilterSummary();
      renderResults();
    }}
    function renderMetrics() {{
      const grid = document.getElementById('metricGrid');
      grid.innerHTML = '';
      payload.metrics.forEach(item => {{
        const btn = document.createElement('button');
        btn.className = 'metric';
        btn.dataset.metric = item.metric;
        const avg = item.avg ? `متوسط: ${{formatPrice(item.avg)}}` : `أسعار معلنة: ${{formatNumber(item.priced)}}`;
        btn.innerHTML = `<span>${{item.label}}</span><strong>${{formatNumber(item.count)}}</strong><small>${{avg}}</small><em>اضغط للاختيار</em>`;
        btn.addEventListener('click', () => applyFilter(item.metric));
        grid.appendChild(btn);
      }});
    }}
    function renderFilters() {{
      populateSelect('governorateFilter', uniqueValues('governorate'), 'كل المحافظات');
      populateSelect('transactionFilter', uniqueValues('transaction'), 'كل المعاملات');
      populateSelect('propertyFilter', uniqueValues('property_type'), 'كل أنواع العقار');
      updateAreaOptions(false);
    }}
    function renderQuickList() {{
      const select = document.getElementById('quickList');
      select.innerHTML = '';
      payload.metrics.forEach(item => {{
        const opt = new Option(`${{item.label}} (${{formatNumber(item.count)}})`, `${{item.metric}}||`);
        select.add(opt);
      }});
      payload.governors.forEach(gov => {{
        [
          ['movement', gov.movement],
          ['sell', gov.sell],
          ['buy', gov.buy],
          ['house', gov.house],
          ['rent', gov.rent],
          ['rent_request', gov.rent_request],
        ].forEach(([metric, count]) => {{
          if (!count) return;
          select.add(new Option(`${{gov.governorate}} - ${{metricLabels[metric]}} (${{formatNumber(count)}})`, `${{metric}}||${{gov.governorate}}`));
        }});
      }});
      select.addEventListener('change', () => {{
        const [metric, governorate] = select.value.split('||');
        applyFilter(metric, governorate);
      }});
    }}
    function renderGovernors() {{
      const tbody = document.querySelector('#govTable tbody');
      tbody.innerHTML = '';
      payload.governors.forEach(gov => {{
        const tr = document.createElement('tr');
        tr.dataset.governorate = gov.governorate;
        tr.innerHTML = `<td>${{gov.governorate}}</td>`;
        [
          ['movement', gov.movement],
          ['sell', gov.sell],
          ['buy', gov.buy],
          ['house', gov.house],
          ['rent', gov.rent],
          ['rent_request', gov.rent_request],
        ].forEach(([metric, count]) => {{
          const td = document.createElement('td');
          td.dataset.metric = metric;
          const btn = document.createElement('button');
          btn.className = 'count-button';
          btn.textContent = formatNumber(count);
          btn.disabled = !count;
          btn.addEventListener('click', () => applyFilter(metric, gov.governorate));
          td.appendChild(btn);
          tr.appendChild(td);
        }});
        tbody.appendChild(tr);
      }});
      highlightGovernorate();
      highlightMetricColumn();
    }}
    function highlightGovernorate() {{
      const gov = selectedGovernorate();
      document.querySelectorAll('#govTable tbody tr').forEach(tr => {{
        tr.classList.toggle('gov-active', !!gov && tr.dataset.governorate === gov);
      }});
    }}
    function highlightMetricColumn() {{
      const table = document.getElementById('govTable');
      table.className = `metric-${{state.metric}}`;
      document.querySelectorAll('#govTable [data-metric]').forEach(cell => {{
        cell.classList.toggle('metric-column-active', cell.dataset.metric === state.metric);
      }});
    }}
    function createFact(label, value) {{
      const box = document.createElement('div');
      box.className = 'fact';
      const b = document.createElement('b');
      b.textContent = label;
      const span = document.createElement('span');
      span.textContent = value || '-';
      box.append(b, span);
      return box;
    }}
    function createDetailsBox(row) {{
      const box = document.createElement('div');
      box.className = 'details-box';
      const title = document.createElement('b');
      title.textContent = 'تفاصيل الإعلان';
      box.appendChild(title);
      const details = row.detailText
        ? [row.detailText]
        : (row.detailTitle ? [row.detailTitle] : [row.summary, row.features].filter(Boolean));
      if (details.length) {{
        details.forEach(text => {{
          const p = document.createElement('p');
          p.textContent = text;
          box.appendChild(p);
        }});
      }} else {{
        const p = document.createElement('p');
        p.className = 'muted';
        p.textContent = 'لا توجد تفاصيل نصية إضافية مسجلة لهذا الإعلان.';
        box.appendChild(p);
      }}
      return box;
    }}
    function renderResultStats(rows) {{
      const priced = rows.filter(row => row.price);
      const avg = priced.length ? priced.reduce((sum, row) => sum + row.price, 0) / priced.length : 0;
      const stats = [
        ['النتائج', formatNumber(rows.length)],
        ['أسعار معلنة', formatNumber(priced.length)],
        ['متوسط السعر', avg ? formatPrice(avg) : 'غير متاح'],
        ['بها صورة', formatNumber(rows.filter(row => row.imageUrl).length)]
      ];
      const container = document.getElementById('resultStats');
      container.innerHTML = '';
      stats.forEach(([label, value]) => {{
        const box = document.createElement('div');
        box.innerHTML = `<b>${{label}}</b><span>${{value}}</span>`;
        container.appendChild(box);
      }});
    }}
    function renderSelectionChips() {{
      const chips = [
        ['المحور', metricLabels[state.metric]],
        ['المحافظة', selectedGovernorate()],
        ['المنطقة', document.getElementById('areaFilter').value],
        ['نوع المعاملة', document.getElementById('transactionFilter').value],
        ['نوع العقار', document.getElementById('propertyFilter').value],
      ];
      const priceText = document.getElementById('priceFilter').selectedOptions[0].textContent;
      if (priceText !== 'كل الأسعار') chips.push(['حالة السعر', priceText]);
      const query = document.getElementById('searchBox').value.trim();
      if (query) chips.push(['بحث', query]);
      const container = document.getElementById('selectionChips');
      container.innerHTML = '';
      chips
        .filter(([, value]) => value)
        .forEach(([label, value]) => {{
          const chip = document.createElement('span');
          chip.textContent = `${{label}}: ${{value}}`;
          container.appendChild(chip);
        }});
    }}
    function currentFilterPairs(includeGovernorate = true) {{
      const pairs = [];
      if (includeGovernorate) pairs.push(['المحافظة', selectedGovernorate()]);
      pairs.push(
        ['نوع المعاملة', document.getElementById('transactionFilter').value],
        ['نوع العقار', document.getElementById('propertyFilter').value],
        ['المنطقة', document.getElementById('areaFilter').value]
      );
      const priceText = document.getElementById('priceFilter').selectedOptions[0].textContent;
      if (priceText !== 'كل الأسعار') pairs.push(['حالة السعر', priceText]);
      const query = document.getElementById('searchBox').value.trim();
      if (query) pairs.push(['بحث', query]);
      return pairs.filter(([, value]) => value);
    }}
    function updateGovernorateFilterSummary() {{
      const target = document.getElementById('govFilterSummary');
      if (!target) return;
      const pairs = currentFilterPairs(false);
      target.textContent = pairs.length ? pairs.map(([label, value]) => `${{label}}: ${{value}}`).join(' | ') : 'كل السجلات';
    }}
    function openDetailsModal(code) {{
      const row = records.find(item => item.code === code);
      if (!row) return;
      const modal = document.getElementById('detailsModal');
      const body = document.getElementById('modalBody');
      document.getElementById('modalTitle').textContent = `${{row.code}} - ${{row.area}}`;
      body.innerHTML = '';
      const imageBox = document.createElement('div');
      imageBox.className = 'modal-image';
      if (row.imageUrl) {{
        const image = document.createElement('img');
        image.src = row.imageUrl;
        image.alt = row.area || row.property_type || 'صورة الإعلان';
        image.referrerPolicy = 'no-referrer';
        image.decoding = 'async';
        image.onerror = () => {{
          imageBox.textContent = 'بدون صورة';
        }};
        imageBox.appendChild(image);
      }} else {{
        imageBox.textContent = 'بدون صورة';
      }}
      const info = document.createElement('div');
      info.className = 'modal-info';
      const badge = document.createElement('span');
      badge.className = 'badge';
      badge.textContent = row.transaction;
      const facts = document.createElement('div');
      facts.className = 'facts';
      facts.append(
        createFact('كود الإعلان', row.code),
        createFact('المحافظة', row.governorate),
        createFact('المنطقة', row.area),
        createFact('نوع العقار', row.property_type),
        createFact('التصنيف', row.detail_class),
        createFact('السعر', row.priceText),
        createFact('المساحة', row.space ? `${{row.space}} م²` : 'غير محدد'),
        createFact('نوع الإعلان', row.listingMode),
        createFact('تاريخ النشر', row.publishedDate || 'غير محدد')
      );
      info.append(badge, facts);
      info.appendChild(createDetailsBox(row));
      const actions = document.createElement('div');
      actions.className = 'modal-actions';
      if (row.originalAvailable && row.originalUrl) {{
        const originalPage = document.createElement('a');
        originalPage.className = 'source';
        originalPage.href = row.originalUrl;
        originalPage.target = '_blank';
        originalPage.rel = 'noopener';
        originalPage.textContent = 'فتح صفحة الإعلان الأصلية';
        actions.appendChild(originalPage);
      }} else {{
        const unavailable = document.createElement('span');
        unavailable.className = 'secondary-source';
        unavailable.textContent = 'صفحة الإعلان الأصلية غير متاحة';
        actions.appendChild(unavailable);
      }}
      info.appendChild(actions);
      body.append(imageBox, info);
      modal.classList.add('show');
    }}
    function closeDetailsModal() {{
      document.getElementById('detailsModal').classList.remove('show');
    }}
    function clearFilters() {{
      state = {{ metric: 'movement', governorate: '', label: metricLabels.movement }};
      ['governorateFilter', 'transactionFilter', 'propertyFilter', 'areaFilter', 'priceFilter', 'sortFilter', 'quickList'].forEach(id => {{
        const el = document.getElementById(id);
        if (!el) return;
        if ('selectedIndex' in el) el.selectedIndex = 0;
        else el.value = '';
      }});
      document.getElementById('searchBox').value = '';
      hideAllSuggestions();
      updateAreaOptions(false);
      document.querySelectorAll('.metric').forEach(btn => {{
        btn.classList.toggle('active', btn.dataset.metric === 'movement');
      }});
      highlightGovernorate();
      highlightMetricColumn();
      renderResults();
    }}
    function renderResults() {{
      const rows = currentRows();
      document.getElementById('resultTitle').textContent = state.label;
      const governorate = selectedGovernorate();
      const activeFilters = [
        governorate,
        document.getElementById('transactionFilter').value,
        document.getElementById('propertyFilter').value,
        document.getElementById('areaFilter').value,
        document.getElementById('priceFilter').selectedOptions[0].textContent !== 'كل الأسعار' ? document.getElementById('priceFilter').selectedOptions[0].textContent : ''
      ].filter(Boolean);
      document.getElementById('resultMeta').textContent = activeFilters.length ? 'النتائج حسب الاختيارات المحددة' : 'اختر من الكروت أو من جدول المحافظات لعرض السجلات';
      document.getElementById('resultCount').textContent = formatNumber(rows.length);
      renderSelectionChips();
      updateGovernorateFilterSummary();
      renderResultStats(rows);
      highlightGovernorate();
      highlightMetricColumn();
      const list = document.getElementById('resultList');
      list.innerHTML = '';
      if (!rows.length) {{
        const empty = document.createElement('div');
        empty.className = 'empty';
        empty.textContent = 'لا توجد سجلات مطابقة للتصفية الحالية.';
        list.appendChild(empty);
        return;
      }}
      rows.forEach(row => {{
        const item = document.createElement('article');
        item.className = 'item';
        const thumb = document.createElement('div');
        thumb.className = 'thumb';
        if (row.imageUrl) {{
          const image = document.createElement('img');
          image.src = row.imageUrl;
          image.alt = row.area || row.property_type || 'صورة الإعلان';
          image.loading = 'lazy';
          image.referrerPolicy = 'no-referrer';
          image.decoding = 'async';
          image.onerror = () => {{
            thumb.textContent = 'بدون صورة';
            thumb.classList.add('no-image');
          }};
          thumb.appendChild(image);
        }} else {{
          thumb.textContent = 'بدون صورة';
        }}
        const body = document.createElement('div');
        body.className = 'item-body';
        const top = document.createElement('div');
        top.className = 'item-top';
        const title = document.createElement('h3');
        title.textContent = `${{row.code}} - ${{row.area}}`;
        const badge = document.createElement('span');
        badge.className = 'badge';
        badge.textContent = row.transaction;
        top.append(title, badge);
        body.appendChild(top);
        const facts = document.createElement('div');
        facts.className = 'facts';
        facts.append(
          createFact('كود الإعلان', row.code),
          createFact('المحافظة', row.governorate),
          createFact('المنطقة', row.area),
          createFact('نوع العقار', row.property_type),
          createFact('السعر', row.priceText),
          createFact('المساحة', row.space ? `${{row.space}} م²` : 'غير محدد')
        );
        body.appendChild(facts);
        const meta = document.createElement('p');
        meta.className = 'desc';
        meta.textContent = `تاريخ النشر: ${{row.publishedDate || 'غير محدد'}} | تفاصيل متاحة: ${{row.detailAvailable}} | صورة: ${{row.hasImage}}`;
        body.appendChild(meta);
        const detailButton = document.createElement('button');
        detailButton.className = 'source';
        detailButton.type = 'button';
        detailButton.textContent = 'عرض التفاصيل';
        detailButton.addEventListener('click', () => openDetailsModal(row.code));
        body.appendChild(detailButton);
        if (row.originalAvailable && row.originalUrl) {{
          const originalPage = document.createElement('a');
          originalPage.className = 'secondary-source';
          originalPage.href = row.originalUrl;
          originalPage.target = '_blank';
          originalPage.rel = 'noopener';
          originalPage.textContent = 'فتح صفحة الإعلان الأصلية';
          body.appendChild(originalPage);
        }} else {{
          const unavailable = document.createElement('span');
          unavailable.className = 'secondary-source';
          unavailable.textContent = 'الأصلية غير متاحة';
          body.appendChild(unavailable);
        }}
        item.append(thumb, body);
        list.appendChild(item);
      }});
    }}
    document.getElementById('searchBox').addEventListener('input', renderResults);
    document.getElementById('priceFilter').addEventListener('change', renderResults);
    document.getElementById('clearFilters').addEventListener('click', clearFilters);
    document.getElementById('printPdfButton').addEventListener('click', () => window.print());
    function handleFilterChange(id) {{
        if (id === 'governorateFilter') {{
          state.governorate = '';
          state.label = metricLabels[state.metric];
          updateAreaOptions(false);
        }}
        if (id === 'transactionFilter') {{
          syncMetricFromTransaction();
        }}
        if (id === 'areaFilter') {{
          syncGovernorateFromArea();
          updateAreaOptions(true);
        }}
        highlightGovernorate();
        highlightMetricColumn();
        renderResults();
    }}
    ['governorateFilter', 'transactionFilter', 'propertyFilter', 'areaFilter', 'sortFilter'].forEach(id => {{
      const el = document.getElementById(id);
      el.addEventListener('change', () => handleFilterChange(id));
      if (searchableFilterIds.includes(id)) {{
        el.addEventListener('input', () => {{
          handleFilterChange(id);
          renderSuggestions(id);
        }});
        el.addEventListener('focus', () => {{
          window.setTimeout(() => el.select(), 0);
          renderSuggestions(id);
        }});
        el.addEventListener('blur', () => window.setTimeout(() => hideSuggestions(id), 150));
        el.addEventListener('keydown', event => {{
          if (event.key === 'Escape') hideSuggestions(id);
          if (event.key === 'Enter') {{
            const first = matchingChoices(id)[0];
            if (first) {{
              event.preventDefault();
              chooseSuggestion(id, first);
            }}
          }}
        }});
      }}
    }});
    document.getElementById('modalClose').addEventListener('click', closeDetailsModal);
    document.getElementById('detailsModal').addEventListener('click', (event) => {{
      if (event.target.id === 'detailsModal') closeDetailsModal();
    }});
    document.addEventListener('keydown', (event) => {{
      if (event.key === 'Escape') closeDetailsModal();
    }});
    renderMetrics();
    renderFilters();
    renderQuickList();
    renderGovernors();
    applyFilter('movement');
  </script>
</body>
</html>
"""
    HTML_PATH.write_text(html, encoding="utf-8")
    PROPERTY_HTML_PATH.write_text(create_property_type_variant(html), encoding="utf-8")


def create_readme(records: list[dict], metrics: list[dict]) -> None:
    lines = [
        "ملحق استعراض الأرقام والعروض الفعلية",
        "",
        "الملفات:",
        "1. 01_لوحة_تفاعلية_للأرقام_والعروض.html: افتحه في Chrome أو Edge، ثم اضغط على أي رقم لعرض السجلات المرتبطة به وعرض التفاصيل داخل اللوحة.",
        "2. 01_لوحة_تفاعلية_للأرقام_والعروض_نسخة_نوع_العقار.html: نسخة بديلة تجعل نوع العقار اختيارا مستقلا، وتعرض جدول المحافظات حسب الحركة فقط.",
        "3. 02_قوائم_العروض_الفعلية_حسب_الأرقام.xlsx: ملف Excel بنفس القوائم مع فلاتر وفرز وروابط صفحة الإعلان الأصلية عند توفرها.",
        "",
        "الأرقام الرئيسية:",
    ]
    lines.extend([f"- {row['label']}: {fmt_int(row['count'])}" for row in metrics])
    lines.extend(
        [
            "",
            f"إجمالي السجلات المعروضة في الملحق: {fmt_int(len(records))}.",
            "يعرض الملحق السجلات المنظفة وتفاصيل الإعلان كما وردت في المصدر عند توفرها.",
        ]
    )
    README_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    records = load_records()
    metrics = metric_summary(records)
    governors = governorate_summary(records)
    remove_detail_pages()
    create_html(records, metrics, governors)
    create_excel(records, metrics, governors)
    update_parent_register_links(records)
    create_readme(records, metrics)
    print(
        json.dumps(
            {
                "output": str(OUTPUT_DIR),
                "records": len(records),
                "metrics": {row["metric"]: row["count"] for row in metrics},
                "files": [HTML_PATH.name, PROPERTY_HTML_PATH.name, XLSX_PATH.name, README_PATH.name],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
